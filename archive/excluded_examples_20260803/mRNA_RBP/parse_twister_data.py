"""
mRNA_RBP/parse_twister_data.py

Parses the real Kobori & Yokobayashi (2016) Twister ribozyme relative-activity
(RA) dataset -- all 144 single and 10,152 double mutants of the Osa-1-4
ribozyme, measured by deep sequencing -- into the pipeline's native
varied-mutrate pool format, so it can be fed straight into
train_twister_deepsquid.py without any oracle-specific special-casing
downstream.

Source: the lab's own Excel file, linked from the paper's Supporting
Information ("The original Excel file is available from
http://yokobayashilab.net/data.html") -- not scraped from the PDF, so full
float precision is preserved. Copied into
mRNA_RBP/data/twister_ribozyme/Kobori_ACIE_2016_Supporting_Data.xlsx.

Sheet "Mutational Matrix" is a 144x144 grid: row/col labels are mutation
codes like "A7U" (WT nt, 1-based position, mutant nt). Diagonal = single-
mutant RA. Off-diagonal (row pos != col pos) = double-mutant RA -- the grid
is symmetric and redundant (M[i,j] == M[j,i]), so each double is collapsed
to one entry (averaged if both triangles are populated).

RA convention: WT itself is not present in the matrix (RA is WT-relative by
definition) -- assigned RA = 1.0 as the anchor training point.

Secondary structure: positions 1-6 are constant flanking sequence (doping
starts at position 7), so only stems fully inside the doped region are
representable here. See CLAUDE.md-adjacent plan notes for the derivation
from the user-supplied dot-bracket structure
"((((((..((((.......((((........)))).....))))....))))))" -- P2 (9-12/41-44)
and P4 (20-23/32-35) are fully inside; P1 (1-6/49-54) is half outside and
excluded. Plus one pseudoknot pair stated explicitly in the paper's main
text: "the Watson-Crick base pair C14-G30" (T2).

Output: mRNA_RBP/outputs_twister_ribozyme/instance_00/varied_mutrate/pool.npz
    nuc_ids    (10297, 48) uint8
    mut_counts (10297,)    uint8  in {0, 1, 2}
    edges      (9, 2)      int32  -- TWISTER_STEM_PAIRS, 0-based array index
    wt_seq     scalar str
    scores_ra  (10297,)    float32
    mut_labels (10297,)    object -- e.g. "WT", "A7U", "A7U,A8G" (traceability)
"""

import os
import re
import sys

import numpy as np
import openpyxl

_HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(_HERE, ".."))

XLSX_PATH = os.path.join(_HERE, "data", "twister_ribozyme",
                         "Kobori_ACIE_2016_Supporting_Data.xlsx")
OUT_PATH = os.path.join(_HERE, "outputs_twister_ribozyme", "instance_00",
                        "varied_mutrate", "pool.npz")

NUC_TO_IDX = {"A": 0, "C": 1, "G": 2, "U": 3}
_WC = {("A", "U"), ("U", "A"), ("C", "G"), ("G", "C")}

_LABEL_RE = re.compile(r"^([ACGU])(\d+)([ACGU])$")

# Paper position numbering (1-based), confirmed against the dot-bracket
# structure + Figure 1A (P2, P4) and the main text (T2 pseudoknot C14-G30).
STEM_PAIRS_PAPER_POS = [
    (9, 44), (10, 43), (11, 42), (12, 41),   # P2
    (20, 35), (21, 34), (22, 33), (23, 32),  # P4
    (14, 30),                                # T2 pseudoknot (main text)
]


def parse_label(label: str):
    m = _LABEL_RE.match(label)
    if not m:
        raise ValueError(f"Unparseable mutation label: {label!r}")
    wt_nuc, pos, mut_nuc = m.group(1), int(m.group(2)), m.group(3)
    return wt_nuc, pos, mut_nuc


def load_matrix(xlsx_path: str):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Mutational Matrix"]
    col_labels = [ws.cell(3, c).value for c in range(2, 146)]
    row_labels = [ws.cell(r, 1).value for r in range(4, 148)]
    assert col_labels == row_labels, "row/col mutation-label mismatch"
    n = len(row_labels)
    assert n == 144, f"expected 144 mutation labels, got {n}"

    M = np.full((n, n), np.nan, dtype=np.float64)
    for ri in range(n):
        for ci in range(n):
            v = ws.cell(4 + ri, 2 + ci).value
            if v is not None:
                M[ri, ci] = float(v)
    return row_labels, M


def derive_wt_seq(labels: list):
    parsed = [parse_label(l) for l in labels]
    positions = sorted({pos for _, pos, _ in parsed})
    pos_to_wt = {}
    for wt_nuc, pos, _ in parsed:
        if pos in pos_to_wt:
            assert pos_to_wt[pos] == wt_nuc, f"inconsistent WT nt at position {pos}"
        else:
            pos_to_wt[pos] = wt_nuc
    lo, hi = positions[0], positions[-1]
    assert list(range(lo, hi + 1)) == positions, "doped positions are not contiguous"
    assert len(positions) * 3 == len(labels), "expected exactly 3 mutation labels per position"
    seq = "".join(pos_to_wt[p] for p in positions)
    return seq, lo, hi


def build_dataset(xlsx_path: str = XLSX_PATH):
    labels, M = load_matrix(xlsx_path)
    wt_seq, pos_lo, pos_hi = derive_wt_seq(labels)
    parsed = [parse_label(l) for l in labels]  # aligned with row/col index
    wt_idx = np.array([NUC_TO_IDX[c] for c in wt_seq], dtype=np.uint8)
    n = len(labels)

    nuc_ids_list  = [wt_idx.copy()]
    scores_list   = [1.0]
    mutcount_list = [0]
    labels_list   = ["WT"]

    for i in range(n):
        _, pos, mut_nuc = parsed[i]
        val = M[i, i]
        assert not np.isnan(val), f"missing diagonal RA for {labels[i]}"
        seq = wt_idx.copy()
        seq[pos - pos_lo] = NUC_TO_IDX[mut_nuc]
        nuc_ids_list.append(seq)
        scores_list.append(val)
        mutcount_list.append(1)
        labels_list.append(labels[i])

    n_doubles = 0
    max_asym = 0.0
    for i in range(n):
        pos_i = parsed[i][1]
        for j in range(i + 1, n):
            pos_j = parsed[j][1]
            if pos_i == pos_j:
                continue
            v1, v2 = M[i, j], M[j, i]
            has1, has2 = not np.isnan(v1), not np.isnan(v2)
            if not has1 and not has2:
                continue
            if has1 and has2:
                max_asym = max(max_asym, abs(v1 - v2))
                val = (v1 + v2) / 2.0
            else:
                val = v1 if has1 else v2
            seq = wt_idx.copy()
            seq[pos_i - pos_lo] = NUC_TO_IDX[parsed[i][2]]
            seq[pos_j - pos_lo] = NUC_TO_IDX[parsed[j][2]]
            nuc_ids_list.append(seq)
            scores_list.append(val)
            mutcount_list.append(2)
            labels_list.append(f"{labels[i]},{labels[j]}")
            n_doubles += 1

    print(f"parsed {n} singles, {n_doubles} doubles "
          f"(max row/col symmetry mismatch: {max_asym:.4f})")
    assert n == 144, f"expected 144 singles, got {n}"
    assert n_doubles == 10_152, f"expected 10152 doubles, got {n_doubles}"

    nuc_ids    = np.stack(nuc_ids_list).astype(np.uint8)
    scores     = np.array(scores_list, dtype=np.float32)
    mut_counts = np.array(mutcount_list, dtype=np.uint8)
    mut_labels = np.array(labels_list, dtype=object)

    total = len(nuc_ids)
    assert total == 1 + 144 + 10_152 == 10_297
    print(f"total sequences: {total}  RA range: [{scores.min():.4f}, {scores.max():.4f}]")

    edges = np.array([(p1 - pos_lo, p2 - pos_lo) for p1, p2 in STEM_PAIRS_PAPER_POS],
                      dtype=np.int32)
    for (i, j), (p1, p2) in zip(edges, STEM_PAIRS_PAPER_POS):
        pair = (wt_seq[i], wt_seq[j])
        assert pair in _WC, (
            f"declared stem pair paper-pos ({p1},{p2}) -> array ({i},{j}) "
            f"is not WC-compatible: wt nts {pair}")
    print(f"stem pairs (paper pos): {STEM_PAIRS_PAPER_POS}")
    print(f"stem pairs (0-based array idx): {[tuple(e) for e in edges]}")
    print(f"WT sequence (positions {pos_lo}-{pos_hi}, L={len(wt_seq)}): {wt_seq}")

    return dict(nuc_ids=nuc_ids, scores=scores, mut_counts=mut_counts,
                mut_labels=mut_labels, wt_seq=wt_seq, edges=edges,
                pos_lo=pos_lo, pos_hi=pos_hi)


def main():
    d = build_dataset()
    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    np.savez_compressed(
        OUT_PATH,
        nuc_ids=d["nuc_ids"],
        mut_counts=d["mut_counts"],
        edges=d["edges"],
        wt_seq=np.array(d["wt_seq"]),
        scores_ra=d["scores"],
        mut_labels=d["mut_labels"],
    )
    print(f"\nSaved -> {OUT_PATH}")


if __name__ == "__main__":
    main()
